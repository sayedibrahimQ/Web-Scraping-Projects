'''
    == Get More Details About Each Offer ==
'''


import basic_data
from requests import get
from bs4 import BeautifulSoup
from openpyxl import Workbook
from re import compile 

W = Workbook()
sheet = W.active
sheet.title = 'Big Data'
DATA = basic_data.DATA
HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

# Sheet Headers
sheet['A1'] = 'CARS_MAKE'
sheet['B1'] = 'MODEL'
sheet['C1'] = 'YEAR'
sheet['D1'] = 'DISTANCE'
sheet['E1'] = 'CONDITION'
sheet['F1'] = 'BODY'
sheet['G1'] = 'ENGINE_SIZE'
sheet['H1'] = 'PRICE'
sheet['I1'] = 'LINK'

for i in range(len(DATA)):
    LINK =  DATA[i][0]
    TITLE =  DATA[i][1]
    PRICE = DATA[i][2]

    # http request for each offer link we got from main page
    source = get(LINK, headers = HEADERS).text
    soup = BeautifulSoup(source, 'lxml')


    # Get all Data we need
    OWNER =  (soup.find('h3', class_='blueColor pointer')).text
    CARS_MAKE = (soup.find('a', class_= 'Make')).text
    MODEL = (soup.find('a', class_= 'Model')).text
    YEAR = (soup.find('a',class_ = 'Year')).text
    DISTANCE = (soup.find('a',class_ = 'Kilometers')).text
    CONDITION = (soup.find('a',class_ = 'Condition')).text
    BODY = (soup.find('a',class_ = 'bold blackColor Body Type')).text
    try:
        ENGINE_SIZE = soup.find('a', class_= 'bold blackColor Engine Size (cc)').text
    except AttributeError:
        ENGINE_SIZE = 'Not Exist'
    DESCRIPTION = ((soup.find('section', id = 'postViewDescription')).find('p', class_= 'inline')).text
    
    # Insert data into execl sheet 
    sheet.cell(column = 1, row = i + 2, value = CARS_MAKE)
    sheet.cell(column = 2, row = i + 2, value = MODEL)
    sheet.cell(column = 3, row = i + 2, value = YEAR)
    sheet.cell(column = 4, row = i + 2, value = DISTANCE)
    sheet.cell(column = 5, row = i + 2, value = CONDITION)
    sheet.cell(column = 6, row = i + 2, value = BODY)
    sheet.cell(column = 7, row = i + 2, value = ENGINE_SIZE)
    sheet.cell(column = 8, row = i + 2, value = PRICE)
    sheet.cell(column = 9, row = i + 2, value = LINK)

W.save('opensouq\\CarsData.xlsx')
